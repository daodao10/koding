#!/usr/bin/env python
# -*- coding: utf-8 -*-

import re
import json

from openpyxl import load_workbook
from openpyxl import Workbook

regPeriod = re.compile(r'\$\("#chart_financial_ratios"\)\.data\("xaxis_categories", \[(.*)\]\);')
regRoe = re.compile(r"name: 'Return On Equity \(ROE\) \[%\]',data: \[(.*)\]")
regNav = re.compile(r"name: 'Net Asset Value \(NAV\) Per Share - Historical \[S\$\]',data: \[(.*)\]")
regEps = re.compile(r"name: 'EPS - Historical \[\$\]',data: \[(.*)\]")
regPe = re.compile(r"name: 'PE Ratio - Historical',data: \[(.*)\]")
regPb = re.compile(r"name: 'Price/NAV - Historical',data: \[(.*)\]")
regPeg = re.compile(r"name: 'Price Earnings to Growth',data: \[(.*)\]")
regDiv = re.compile(r"name: 'Dividend Yield - Historical \[%\]',data: \[(.*)\]")

def getFAContent(symbol):
    try:
        f = open("fa/%s.txt" % symbol, mode='r')
        content = f.read()
        f.close()       
        if len(content) > 0:
            return content
    except Exception, e:
        pass
    return None

def getValue(reg, content):
    result = reg.search(content)
    if result and result.lastindex > 0:
        return result.group(1)
    
    return None

def parseFA(content):
    result = {}
    period = getValue(regPeriod, content)

    if period:
        # period
        period = ','.join(re.findall(r'\d{4}', period))
        result["year"] = period

        # roe
        roe = getValue(regRoe, content)
        if roe:
            result["roe"] = roe

        # nav
        nav = getValue(regNav, content)
        if nav:
            result["nav"] = nav

        # eps
        eps = getValue(regEps, content)
        if eps:
            result["eps"] = eps

        # pe
        pe = getValue(regPe, content)
        if pe:
            result["pe"] = pe

        # pb
        pb = getValue(regPb, content)
        if pb:
            result["pb"] = pb

        # peg
        peg = getValue(regPeg, content)
        if peg:
            result["peg"] = peg

        # div
        div = getValue(regDiv, content)
        if div:
            result["div"] = div

    return result

def generateXlsx(data):
    faFile = "test.xlsx"
    wb = Workbook()
    # wb = load_workbook(faFile)
    i = 0
    year = 1994
    for k, v in data.items():
        if k != "year":
            # print "key: %s, value: %s" % (k, v)
            wb.create_sheet(i, k)
            for col in range(0, 21):
               ws = wb.get_sheet_by_name(k)
               ws.cell(row = 1, column = 2 + col).value = 1994 + col
            i += 1

    # ws = wb.get_sheet_by_name('eps')
    # ws = wb.active
    # maxRow = ws.get_highest_row()
    # # ws.cell(row, col).value = 

    wb.save(faFile)

def populate(symbol, data):
    faFile = "test.xlsx"
    wb = load_workbook(faFile)

    baseCol = int(data["year"][:4]) - 1994 + 2
    count = len(data["year"].split(','))

    # fill sheets
    for k, v in data.items():
        if k != "year":
            values = v.split(',')
            ws = wb.get_sheet_by_name(k)
            baseRow = ws.get_highest_row() + 1
            
            ws.cell(row = baseRow, column = 1).value = symbol

            # fill columns
            for i in range(0, count):
                ws.cell(row = baseRow, column = baseCol + i).value = values[i]

    wb.save(faFile)

# roa, 
# data = {"year": "2005,2006,2007,2008,2009,2010,2011,2012,2013,2014", 
#         "eps": "0.0457,0.0976,0.172,0.1773,0.13623,0.07,0.05,0.07,0.1162,0.0511", 
#         "roe": "11.676,21.489,31.02,24.066,17.708,8.878,5.703,8.577,10.361,4.933", 
#         "pb": "1.0974,1.4022,2.0618,3.1689,1.5774,1.1535,1.05,0.8082,1.1491,0.9849", 
#         "nav": "0.39,0.45,0.55,0.74,0.7709,0.86,0.9,0.98,1.08,1.06", 
#         "pe": "9.37,6.47,6.59,13.23,8.93,14.17,18.9,11.31,10.68,20.43", 
#         "div": "1.869,4.754,3.968,3.241,4.112,2.52,1.587,3.157,3.223,1.916", 
#         "peg": "0.196,0.056,0.084,0.964,null,null,null,0.18,0.161,null"
# }
# generateXlsx(data)


symbol = "AAJ.SI"
content = getFAContent(symbol)
if content:
    data = parseFA(content)
    print data
    # populate(counter, data)
else:
    print '%s error' % counter

print 'done'